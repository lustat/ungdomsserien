import pandas as pd
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color


def adjust_column_width(ws, df: pd.DataFrame, default_width=10):
    dim_holder = DimensionHolder(worksheet=ws)
    for (df_col, col) in zip(df.columns, ws.columns):
        header_cell = col[0]
        header_cell.font = Font(bold=True)
        if isinstance(df[df_col].iloc[0], str):
            adjusted_width = 1.3 * df[df_col].str.len().max()
            adjusted_width = min(adjusted_width, 40)
            adjusted_width = max(adjusted_width, 8)
        elif isinstance(df[df_col].iloc[0], int) | isinstance(df[df_col].iloc[0], float):
            adjusted_width = default_width
        else:
            adjusted_width = default_width
        dim_holder[get_column_letter(col[0].col_idx)] = ColumnDimension(ws, min=col[0].col_idx, max=col[0].col_idx,
                                                                        width=adjusted_width)

    ws.column_dimensions = dim_holder

    # Freeze panes
    c = ws['B2']
    ws.freeze_panes = c
    return ws
