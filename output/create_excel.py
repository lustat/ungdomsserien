import pandas as pd
import os
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color

from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from output.excel_helpers import adjust_column_width


def individual_results_excel(storage_path, dct):
    if dct:  # if dictionary is non-empty
        # One Excel sheet per class
        today = datetime.today().strftime('%y%m%d')
        excel_name = 'IndividualResults_' + today + '.xlsx'
        excel_file = os.path.join(storage_path, excel_name)

        wb = openpyxl.Workbook()
        standard_sheets = wb.get_sheet_names()
        for class_name in dct.keys():
            worksheet = wb.create_sheet(class_name)

            df = dct[class_name]
            if not df.empty:
                if sum(df.night) == 0:
                    df = df.drop(columns=['night'])

                if all(df.score == df.total):
                    df = df.drop(columns=['score'])

                if 'position' in df.columns:
                    new_column_order = ['position']
                    columns = list(df.columns)
                    columns.remove('position')
                    new_column_order.extend(columns)
                    df = df[new_column_order]

                for r in dataframe_to_rows(df, index=False, header=True):
                    worksheet.append(r)
                worksheet = adjust_column_width(worksheet, df)

        # Remove standard sheets
        for standard_sheet in standard_sheets:
            std = wb.get_sheet_by_name(standard_sheet)
            wb.remove(std)

        wb.save(excel_file)
        print('Sparar ' + excel_file)
        return excel_file


def club_results_to_excel(storage_path, df, club_results):
    def add_line_flag(data):
        data = data.assign(position=0)
        data = data.assign(negative_position=0)
        for (key, row) in data.iterrows():
            division_data = data.loc[data.division == row.division]
            data.at[key, 'position'] = 1 + sum(row.score < division_data.score)
            data.at[key, 'negative_position'] = 1 + sum(row.score > division_data.score)

        data = data.assign(line_above=False)
        data = data.assign(red_place=False)
        data = data.assign(green_place=False)

        for division in ['Elit', 'Division 1', 'Division 2', 'Division 3']:
            division_data = data.loc[data.division == division]
            if len(division_data) >= 4:
                first_or_2nd = division_data.loc[(division_data.position == 1) |
                                                 (division_data.position == 2)]
                if len(first_or_2nd) == 2:
                    if division.lower() != 'elit':
                        data.at[first_or_2nd.index[0], 'green_place'] = True
                        data.at[first_or_2nd.index[1], 'green_place'] = True

                third_position = division_data.loc[division_data.position == 3]
                if len(third_position) == 1:
                    if division.lower() != 'elit':
                        data.at[third_position.index[0], 'line_above'] = True

                second_last = division_data.loc[division_data.negative_position == 2]
                if len(second_last) == 1:
                    if division.lower() != 'division 3':
                        data.at[second_last.index[0], 'line_above'] = True

                last_or_2nd_last = division_data.loc[(division_data.negative_position == 1) |
                                                     (division_data.negative_position == 2)]
                if len(last_or_2nd_last) == 2:
                    if division.lower() != 'division 3':
                        data.at[last_or_2nd_last.index[0], 'red_place'] = True
                        data.at[last_or_2nd_last.index[1], 'red_place'] = True

        data = data.drop(columns={'position', 'negative_position'})
        return data

    def create_summary_with_division(worksheet, data):
        previous_division = ''
        last_row_header = True
        for row in dataframe_to_rows(data, index=False, header=True):
            current_division = row[-4]
            line_above = row[-3]
            red_place = row[-2]
            green_place = row[-1]
            print_row = row[:-4]
            if previous_division:
                if current_division == previous_division:
                    worksheet.append(print_row)
                    # if line_above:
                    #     for column_number in range(1, len(print_row) + 1):
                    #         cell = worksheet.cell(worksheet.max_row, column_number)
                    #         cell.border = Border(top=Side(style='dotted'))
                    if red_place:
                        for column_number in range(1, len(print_row) + 1):
                            cell = worksheet.cell(worksheet.max_row, column_number)
                            cell.fill = PatternFill('lightTrellis', fgColor=Color('55FF8888'))
                    if green_place:
                        for column_number in range(1, len(print_row) + 1):
                            cell = worksheet.cell(worksheet.max_row, column_number)
                            cell.fill = PatternFill('lightTrellis', fgColor=Color('5588FF88'))
                else:
                    empty_row = [' ' for str in row]
                    if last_row_header:
                        worksheet.append(empty_row)
                        last_row_header = False
                    else:
                        worksheet.append(empty_row)
                        worksheet.append(empty_row)
                    title_row = empty_row
                    title_row[0] = current_division
                    worksheet.append(title_row)
                    cell = worksheet.cell(worksheet.max_row, 1)
                    cell.font = Font(bold=True)
                    worksheet.append(print_row)
                    if green_place:
                        for column_number in range(1, len(print_row) + 1):
                            cell = worksheet.cell(worksheet.max_row, column_number)
                            cell.fill = PatternFill('lightTrellis', fgColor=Color('5588FF88'))
            else:
                worksheet.append(print_row)
            previous_division = current_division

    def create_summary_without_division(worksheet, data):
        for row in dataframe_to_rows(data, index=False, header=True):
            worksheet.append(row[:-1])

    if df.empty:
        return None
    else:
        today = datetime.today().strftime('%y%m%d')
        excel_name = 'ClubResults_' + today + '.xlsx'
        excel_file = os.path.join(storage_path, excel_name)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        df = df.reset_index(drop=True, inplace=False)

        if all(df.division.isna()):
            create_summary_without_division(ws, df)
        else:
            df = add_line_flag(df)
            create_summary_with_division(ws, df)

        ws = adjust_column_width(ws, df)

        for club_result in club_results:
            if not club_result.empty:
                if 'name' in club_result.columns:
                    name_first_list = ['name']
                    name_first_list.extend([col for col in club_result.columns if col != 'name'])
                    club_result = club_result[name_first_list]

                do_not_print_columns = ['event_year', 'region', 'birthyear', 'seconds']
                club_result = club_result.drop(columns=do_not_print_columns)

                ws = wb.create_sheet(club_result.iloc[0].club)
                for r in dataframe_to_rows(club_result, index=False, header=True):
                    ws.append(r)

                ws = adjust_column_width(ws, club_result)

        print('Sparar ' + excel_file)

        wb.save(excel_file)
        wb.close()
        return excel_file

