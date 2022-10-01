import pandas as pd
import os
import numpy as np
from openpyxl import load_workbook

from loader.input_structure import get_input_structure


def get_excel_sheets(excel_file: str):
    wb = load_workbook(excel_file)
    return wb.sheetnames


def check_sheet_names(sheet_names):
    ok_sheets = []

    for original_sheet in sheet_names:
        sheet = original_sheet.replace(' ', '')
        if sheet.isdigit():
            ok_sheets.append(original_sheet)
        else:
            if (sheet.lower().startswith('night')) | (sheet.lower().startswith('natt')):
                ok_sheets.append(original_sheet)
            elif sheet.lower().startswith('division'):
                ok_sheets.append(original_sheet)
            elif sheet.lower().startswith('parameters'):
                ok_sheets.append(original_sheet)
    return ok_sheets


def check_columns_in_sheets(excel_file, sheets):
    day_columns = ['name', 'class', 'club']
    night_columns = ['personid', 'name', 'club', 'useries', 'nightclass', 'eventid', 'finished']
    division_columns = ['division', 'clubid', 'club']
    parameter_columns = ['parameter', 'value']

    filename = os.path.split(excel_file)[1]
    column_missing = False
    for sheet in sheets:
        df = pd.read_excel(excel_file, sheet)
        current_columns = [col.lower().replace(' ', '').replace('-', '') for col in df.columns]
        if sheet.replace(' ', '').isdigit():
            for column in day_columns:
                if not (column in current_columns):
                    print('Varning: ' + column + ' saknas i ' + sheet)
                    column_missing = True
        else:
            if (sheet.lower().startswith('night')) | (sheet.lower().startswith('natt')):
                for column in night_columns:
                    if not (column in current_columns):
                        print('Varning: ' + column + ' saknas i ' + sheet)
                        column_missing = True
            if sheet.lower().startswith('division'):
                for column in division_columns:
                    if column not in current_columns:
                        print('Varning: ' + column + ' saknas i ' + sheet)
                        column_missing = True
            if sheet.lower().startswith('parameters'):
                for column in parameter_columns:
                    if column not in current_columns:
                        print('Varning: ' + column + ' saknas i ' + sheet)
                        column_missing = True

    if column_missing:
        print('Excel-fil:' + filename + ' saknar data')


def check_excel_input(manual_input_file):
    print('Granskar ' + manual_input_file)
    template = get_input_structure()
    template_sheets_lower = [sheet.lower() for sheet in template.keys()]
    template_digit_sheets = [sheet for sheet in template.keys() if sheet.isdigit()]
    if not template_digit_sheets:
        raise ValueError('Template must contain at least one sheet with a digit name')
    template_digit_sheet = template_digit_sheets[0]

    all_sheets = get_excel_sheets(manual_input_file)
    all_sheets_lower = [sheet.lower() for sheet in all_sheets]

    necessary_sheets = [sheet for sheet in template.keys() if template[sheet]['compulsory']]
    for necessary_sheet in necessary_sheets:
        if necessary_sheet.lower() not in all_sheets_lower:
            raise ValueError(necessary_sheet.capitalize() + '-flik saknas i ' + manual_input_file)

    valid_sheets = []
    neglected_sheets = []
    for sheet in all_sheets:
        if sheet.replace(' ', '').isdigit():
            valid_sheets.append(sheet)
        else:
            if sheet.lower() in template_sheets_lower:
                valid_sheets.append(sheet)
            else:
                neglected_sheets.append(sheet)

    for sheet in valid_sheets:
        if sheet.replace(' ', '').isdigit():
            template_df = template[template_digit_sheet]['example_df']
        else:
            template_df = template[sheet.lower()]['example_df']

        input_df = pd.read_excel(manual_input_file, sheet)
        for template_column in template_df.columns:
            if template_column not in input_df.columns:
                raise ValueError('Kolumnen ' + template_column + ' saknas i fliken ' + sheet)

        if sheet.lower().replace(' ', '') == 'parameters':  # Special check of Parameters-sheet
            input_df.columns = [col.lower() for col in input_df.columns]
            for parameter in template_df['Parameter']:
                if parameter not in list(input_df['parameter']):
                    raise ValueError('Raden med ' + parameter + ' saknas i fliken ' + sheet)
                else:
                    value = input_df.loc[input_df['parameter'] == parameter, 'value'].values[0]
                    print(parameter + ' = ' + str(value))

        print('Flik "' + sheet + '": OK')

    if neglected_sheets:
        for neglected_sheet in neglected_sheets:
            print('Flik "' + neglected_sheet + '" läses ej in')

    return valid_sheets


def read_manual_input(manual_input_file='C:\\Users\\Klas\\Desktop\\Example_inputs\\Manual_input.xlsx'):
    race_to_manual_input = {}
    division_table = pd.DataFrame()
    if not os.path.exists(manual_input_file):
        if manual_input_file:
            raise ValueError('Ingen Excel-fil identifierad med namn: ' + manual_input_file)
        else:
            raise ValueError('Ingen Excel-fil vald')
    else:
        accepted_sheets = check_excel_input(manual_input_file)

        for sheet in accepted_sheets:
            df = pd.read_excel(manual_input_file, sheet)
            sheet = sheet.replace(' ', '').lower()
            df.columns = [col.lower().replace(' ', '').replace('-', '') for col in df.columns]
            if sheet.isdigit():
                race_to_manual_input[int(sheet)] = df
            else:
                if sheet.startswith('division'):
                    division_table = df
                elif sheet.startswith('night'):
                    race_to_manual_input['night'] = df
                elif sheet.startswith('parameters'):
                    user_input = df
                    user_input = user_input.set_index('parameter')
                    user_input = user_input.to_dict()['value']
                    if 'night_ids' in user_input.keys():
                        if isinstance(user_input['night_ids'], float):
                            if np.isnan(user_input['night_ids']):
                                user_input.pop('night_ids', None)
                    if 'event_ids' not in user_input.keys():
                        raise ImportError('Parameter "event_ids" saknas i Excel-fil. Lägg till i fliken "Parameters"')
                    else:
                        if isinstance(user_input['event_ids'], float):
                            if np.isnan(user_input['event_ids']):
                                raise ImportError(
                                    'Parameter "event_ids" måste fyllas i ' + manual_input_file)

    return race_to_manual_input, division_table, user_input


if __name__ == '__main__':
    test_excel_folder = 'C:\\Users\\Klas\\Desktop\\Example_inputs'
    excel_names = [file for file in os.listdir(test_excel_folder) if file.endswith('.xlsx')]
    excel_files = [os.path.join(test_excel_folder, file_name) for file_name in excel_names]
    for file in excel_files:
        try:
            dct, div_df, user_dct = read_manual_input(file)
        except ValueError:
            print(file + ' is incorrect')
            print('  ')
            print('  ')
        else:
            print(file + ' is correct')
            print('  ')
            print('  ')
    print('Finished')