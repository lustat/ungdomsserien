import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color


def get_input_structure():
    event_ids = [20550, 21406, 21376, 21988, 21732, 21644]
    df_parameters = pd.DataFrame(data={'Parameter': ['event_ids', 'night_ids'],
                                       'Value': ['20550, 21406, 21376, 21988, 21732, 21644', '21851, 21961']})

    dct_parameters = {'example_df': df_parameters,
                      'compulsory': True,
                      'description': 'Listar input-parametrar till beräkning (t ex tävlingars Event ID)'}

    division = list(np.tile(['Elit'], 6))
    division.extend(list(np.tile(['Division 1'], 7)))
    division.extend(list(np.tile(['Division 2'], 7)))
    division.extend(list(np.tile(['Division 3'], 4)))
    club_id = [483, 639, 125, 258, 471, 114, 35, 614, 54, 487, 167, 169,
               507, 165, 81, 111, 294, 461, 400, 558, 132, 331, 544, 228]

    club_name = ['Lunds OK', 'Skåneslättens OL', 'Helsingborgs SOK', 'Malmö OK', 'Örkelljunga FK', 'FK Göingarna',
                 'Andrarums IF', 'Hjärnarps OL', 'FK Boken', 'OK Tyringe', 'Hässleholms OK', 'Hästveda OK',
                 'OK Kontinent', 'Härlövs IF', 'Frosta OK', 'OK Gynge', 'OK Pan-Kristianstad', 'FK Åsen',
                 'Tormestorps IF', 'OK Kompassen', 'Eslövs FK', 'IS Skanne', 'Ystads OK', 'IS Kullen']

    df_division = pd.DataFrame(data={'Division': division,
                                     'Clubid': club_id,
                                     'Club': club_name})

    dct_division = {'example_df': df_division,
                    'compulsory': False,
                    'description': 'Anger divisionstillhörighet for respektive klubb'}

    df_event = pd.DataFrame(data={'Name': ['Förnamn1 Efternamn1', 'Förnamn2 Efternamn2'], 'Class': ['U1', 'Inskolning'],
                                  'Club': ['Skåneslättens OL', 'Lunds OK']})
    dct_event = {'example_df': df_event,
                 'compulsory': False,
                 'description': 'Anger namn på tävlande med okänd ålder som ska räknas'}

    df_night = pd.DataFrame(data={'Person ID': ['11111', '222222'],
                                  'Name': ['Förnamn1 Efternamn1', 'Förnamn2 Efternamn2'],
                                  'Club': ['Skåneslättens OL', 'Lunds OK'],
                                  'U-series': ['D12', 'H14'],
                                  'Night-class': ['U2', 'Öppen Motion 5'],
                                  'Event ID': [21930, 22921],
                                  'Finished': [1, 0]})
    dct_night = {'example_df': df_night,
                 'compulsory': False,
                 'description': 'Anger deltagande i natt-tävling, i natt-tävlingar som inte listats via "night_ids" i Parameters-flik.'}

    structure = {'parameters': dct_parameters,
                 'divisions': dct_division,
                 str(event_ids[0]): dct_event,
                 'night': dct_night}
    return structure


def create_excel_template(structure, template_path=''):
    def adjust_columns(max_column_width=40):
        for (df_col, col) in zip(df.columns, worksheet.columns):
            column = col[0].column
            header_cell = col[0]
            header_cell.font = Font(bold=True)
            if isinstance(df[df_col].iloc[0], str):
                adjusted_width = 1.4 * df[df_col].str.len().max()
                adjusted_width = min(adjusted_width, max_column_width)
                adjusted_width = max(adjusted_width, 10)
            elif isinstance(df[df_col].iloc[0], int) | isinstance(df[df_col].iloc[0], float):
                adjusted_width = 10
            else:
                adjusted_width = 10
            worksheet.column_dimensions[column].width = adjusted_width

    if not template_path:
        print('Debug-mode in ' + __name__)
        template_path = os.path.split(os.getcwd())[0]

    excel_name = 'Input_Example.xlsx'
    excel_file = os.path.join(template_path, excel_name)

    wb = openpyxl.Workbook()
    standard_sheets = wb.get_sheet_names()
    for sheet in structure.keys():
        worksheet = wb.create_sheet(sheet.capitalize())
        df = structure[sheet]['example_df']

        for r in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(r)
        adjust_columns()

    # Add description sheet
    worksheet = wb.create_sheet('Description')
    df = pd.DataFrame()
    for sheet in structure.keys():
        if 'description' in structure[sheet].keys():
            df.at[sheet, 'Sheet'] = sheet.capitalize()
            df.at[sheet, 'Comment'] = structure[sheet]['description']

    for r in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(r)
    adjust_columns(max_column_width=100)

    # Remove standard sheets
    for standard_sheet in standard_sheets:
        std = wb.get_sheet_by_name(standard_sheet)
        wb.remove(std)

    wb.save(excel_file)
    print('Skapar exempel-fil enligt mall: ' + excel_file)
    return excel_file


if __name__ == '__main__':
    template = get_input_structure()
    file = create_excel_template(structure=template)

    print('Finished')